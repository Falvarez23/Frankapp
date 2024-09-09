import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Título de la aplicación
st.title("Formulario de Datos")

# Crear el formulario
with st.form("my_form"):
    nombre = st.text_input("Nombre:")
    email = st.text_input("Correo electrónico:")
    edad = st.number_input("Edad:", min_value=0, max_value=120)
    
    # Enviar los datos
    submitted = st.form_submit_button("Enviar")

# Cuando se envían los datos
if submitted:
    # Crear un DataFrame con los datos
    datos = pd.DataFrame({
        "Nombre": [nombre],
        "Correo": [email],
        "Edad": [edad]
    })

    # Guardar los datos en un archivo Excel
    try:
        # Si el archivo ya existe, cargarlo y agregar datos
        book = load_workbook("datos.xlsx")
        writer = pd.ExcelWriter("datos.xlsx", engine="openpyxl")
        writer.book = book
        datos.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        writer.save()
        writer.close()
    except FileNotFoundError:
        # Si el archivo no existe, crearlo
        datos.to_excel("datos.xlsx", index=False)

    st.success("¡Datos guardados exitosamente!")

    # Mostrar los datos en la app
    st.dataframe(datos)
